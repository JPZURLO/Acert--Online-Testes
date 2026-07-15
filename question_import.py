import io
import re
import unicodedata
import uuid
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


MAX_IMPORTED_QUESTIONS = 200
MAX_IMPORT_FILE_BYTES = 5 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 50 * 1024 * 1024
ALTERNATIVE_PATTERN = re.compile(r"^alternativa\s+([a-j])$")


class QuestionImportError(ValueError):
    def __init__(self, message, errors=None):
        super().__init__(message)
        self.errors = errors or []


def normalize_text(value):
    text = str(value if value is not None else "").strip().lower()
    return "".join(
        character
        for character in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(character)
    )


def display_text(value, maximum=3000):
    return str(value if value is not None else "").strip()[:maximum]


def question_type(value):
    normalized = normalize_text(value).replace("_", " ").replace("-", " ")
    aliases = {
        "multipla escolha": "multiple_choice",
        "multiple choice": "multiple_choice",
        "objetiva": "multiple_choice",
        "verdadeiro ou falso": "true_false",
        "verdadeiro falso": "true_false",
        "true false": "true_false",
        "dissertativa": "essay",
        "discursiva": "essay",
        "essay": "essay",
    }
    return aliases.get(normalized)


def integer_value(value, row_number, errors):
    if value in (None, ""):
        return 10
    try:
        number = float(value)
        if not number.is_integer() or not 0 <= number <= 1000:
            raise ValueError
        return int(number)
    except (TypeError, ValueError):
        errors.append(f"Linha {row_number}: Pontos deve ser um número inteiro entre 0 e 1000.")
        return 10


def boolean_value(value, row_number, errors):
    if value in (None, ""):
        return True
    normalized = normalize_text(value)
    if normalized in {"sim", "s", "true", "verdadeiro", "1"}:
        return True
    if normalized in {"nao", "n", "false", "falso", "0"}:
        return False
    errors.append(f"Linha {row_number}: Obrigatória deve conter Sim ou Não.")
    return True


def correct_multiple_choice(value, options, row_number, errors):
    answer = display_text(value, 500)
    normalized = normalize_text(answer)
    letter_match = re.fullmatch(r"(?:alternativa\s+)?([a-j])", normalized)
    if letter_match:
        option_index = ord(letter_match.group(1)) - ord("a")
        if option_index < len(options):
            return options[option_index]
    for option in options:
        if normalize_text(option) == normalized:
            return option
    errors.append(
        f"Linha {row_number}: Resposta correta deve ser a letra ou o texto exato de uma alternativa preenchida."
    )
    return ""


def correct_true_false(value, row_number, errors):
    normalized = normalize_text(value)
    if normalized in {"verdadeiro", "v", "true"}:
        return "Verdadeiro"
    if normalized in {"falso", "f", "false"}:
        return "Falso"
    errors.append(f"Linha {row_number}: Resposta correta deve ser Verdadeiro ou Falso.")
    return ""


def locate_headers(sheet):
    for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        columns = {}
        alternatives = []
        for column_index, value in enumerate(row):
            normalized = normalize_text(value)
            if normalized == "tipo":
                columns["type"] = column_index
            elif normalized in {"enunciado", "pergunta", "questao"}:
                columns["prompt"] = column_index
            elif normalized in {"pontos", "pontuacao"}:
                columns["points"] = column_index
            elif normalized in {"obrigatoria", "obrigatorio"}:
                columns["required"] = column_index
            elif normalized in {"resposta correta", "gabarito"}:
                columns["correct"] = column_index
            else:
                match = ALTERNATIVE_PATTERN.fullmatch(normalized)
                if match:
                    alternatives.append((match.group(1), column_index))
        if {"type", "prompt"}.issubset(columns):
            columns["alternatives"] = [index for _, index in sorted(alternatives)]
            return row_number, columns
    raise QuestionImportError(
        "Não foi possível localizar as colunas Tipo e Enunciado. Use o modelo disponível no site."
    )


def cell(row, column_index):
    if column_index is None or column_index >= len(row):
        return None
    return row[column_index]


def parse_question_workbook(stream):
    payload = stream.read(MAX_IMPORT_FILE_BYTES + 1)
    if len(payload) > MAX_IMPORT_FILE_BYTES:
        raise QuestionImportError("O arquivo deve ter no máximo 5 MB.")
    try:
        with ZipFile(io.BytesIO(payload)) as archive:
            entries = archive.infolist()
            if len(entries) > 1000 or sum(entry.file_size for entry in entries) > MAX_UNCOMPRESSED_BYTES:
                raise QuestionImportError("O arquivo Excel é muito complexo para importação.")
    except BadZipFile as exc:
        raise QuestionImportError("O arquivo não é um Excel .xlsx válido.") from exc

    try:
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    except (InvalidFileException, BadZipFile, OSError, ValueError) as exc:
        raise QuestionImportError("O arquivo não é um Excel .xlsx válido.") from exc

    try:
        sheet = next(
            (workbook[name] for name in workbook.sheetnames if normalize_text(name) == "questoes"),
            workbook.active,
        )
        header_row, columns = locate_headers(sheet)
        questions = []
        errors = []

        for row_number, row in enumerate(
            sheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            if all(value in (None, "") for value in row):
                continue
            if len(questions) >= MAX_IMPORTED_QUESTIONS:
                errors.append(f"O arquivo ultrapassa o limite de {MAX_IMPORTED_QUESTIONS} questões.")
                break

            prompt = display_text(cell(row, columns.get("prompt")))
            imported_type = question_type(cell(row, columns.get("type")))
            if not prompt:
                errors.append(f"Linha {row_number}: informe o enunciado da questão.")
            if not imported_type:
                errors.append(
                    f"Linha {row_number}: Tipo deve ser Múltipla escolha, Verdadeiro ou falso ou Dissertativa."
                )

            options = [
                display_text(cell(row, column_index), 500)
                for column_index in columns.get("alternatives", [])
            ]
            options = [option for option in options if option]
            correct_value = cell(row, columns.get("correct"))

            if imported_type == "multiple_choice":
                if len(options) < 2:
                    errors.append(f"Linha {row_number}: preencha pelo menos as alternativas A e B.")
                correct_answer = correct_multiple_choice(correct_value, options, row_number, errors)
            elif imported_type == "true_false":
                options = ["Verdadeiro", "Falso"]
                correct_answer = correct_true_false(correct_value, row_number, errors)
            else:
                options = []
                correct_answer = display_text(correct_value, 500)

            questions.append(
                {
                    "id": f"imported-{uuid.uuid4()}",
                    "type": imported_type or "multiple_choice",
                    "prompt": prompt,
                    "points": integer_value(cell(row, columns.get("points")), row_number, errors),
                    "required": boolean_value(cell(row, columns.get("required")), row_number, errors),
                    "options": options,
                    "correctAnswer": correct_answer,
                }
            )

        if not questions:
            raise QuestionImportError("Nenhuma questão preenchida foi encontrada na aba Questões.")
        if errors:
            raise QuestionImportError(
                "Corrija os itens indicados no Excel e tente novamente.",
                errors[:25],
            )
        return questions
    finally:
        workbook.close()
