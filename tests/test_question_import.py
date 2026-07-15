import io
import unittest
from pathlib import Path

from flask import Flask, jsonify
from openpyxl import Workbook, load_workbook

from company_api import create_company_blueprint
from question_import import QuestionImportError, parse_question_workbook


HEADERS = [
    "Tipo",
    "Enunciado",
    "Pontos",
    "Obrigatória",
    "Alternativa A",
    "Alternativa B",
    "Alternativa C",
    "Resposta correta",
]


def workbook_bytes(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Questões"
    sheet.append(HEADERS)
    for row in rows:
        sheet.append(row)
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


class QuestionImportTests(unittest.TestCase):
    def test_valid_workbook_normalizes_all_question_types(self):
        stream = workbook_bytes(
            [
                ["Múltipla escolha", "Pergunta objetiva", 20, "Sim", "Certa", "Errada", "", "A"],
                ["Verdadeiro ou falso", "Afirmação", 10, "Não", "", "", "", "Falso"],
                ["Dissertativa", "Explique o processo", 30, "Sim", "", "", "", "Critério esperado"],
            ]
        )
        questions = parse_question_workbook(stream)
        self.assertEqual([question["type"] for question in questions], ["multiple_choice", "true_false", "essay"])
        self.assertEqual(questions[0]["correctAnswer"], "Certa")
        self.assertEqual(questions[1]["options"], ["Verdadeiro", "Falso"])
        self.assertFalse(questions[1]["required"])
        self.assertEqual(questions[2]["points"], 30)

    def test_invalid_rows_return_line_specific_errors(self):
        stream = workbook_bytes([["Múltipla escolha", "Sem alternativas", 10, "Talvez", "", "", "", "Z"]])
        with self.assertRaises(QuestionImportError) as context:
            parse_question_workbook(stream)
        self.assertTrue(any("Linha 2" in error for error in context.exception.errors))
        self.assertTrue(any("alternativas A e B" in error for error in context.exception.errors))

    def test_authenticated_endpoint_parses_excel_without_opening_database(self):
        app = Flask(__name__)

        def company_session(_expected_type):
            return {"sub": "7"}, None

        def should_not_open_database():
            raise AssertionError("A validação do Excel não deve abrir o banco")

        app.register_blueprint(create_company_blueprint(should_not_open_database, company_session))
        response = app.test_client().post(
            "/api/company/question-imports",
            data={
                "file": (
                    workbook_bytes([["Verdadeiro ou falso", "Afirmação", 15, "Sim", "", "", "", "Verdadeiro"]]),
                    "questoes.xlsx",
                )
            },
            content_type="multipart/form-data",
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["count"], 1)
        self.assertEqual(response.get_json()["totalPoints"], 15)

    def test_import_endpoint_requires_company_session(self):
        app = Flask(__name__)

        def reject_session(_expected_type):
            return None, (jsonify({"message": "Acesso não autorizado."}), 403)

        app.register_blueprint(create_company_blueprint(lambda: None, reject_session))
        self.assertEqual(app.test_client().post("/api/company/question-imports", data={}).status_code, 403)

    def test_downloadable_template_matches_the_import_contract(self):
        template = Path("front-end/assets/templates/modelo-importacao-questoes.xlsx")
        self.assertTrue(template.exists())
        workbook = load_workbook(template, read_only=True, data_only=True)
        try:
            self.assertIn("Questões", workbook.sheetnames)
            headers = [cell.value for cell in next(workbook["Questões"].iter_rows(min_row=4, max_row=4))]
            self.assertEqual(headers[0:4], ["Tipo", "Enunciado", "Pontos", "Obrigatória"])
            self.assertEqual(headers[-1], "Resposta correta")
        finally:
            workbook.close()

    def test_editor_exposes_download_and_upload_controls(self):
        html = Path("front-end/login_cliente.html").read_text(encoding="utf-8")
        script = Path("front-end/js/company-dashboard.js").read_text(encoding="utf-8")
        self.assertIn("modelo-importacao-questoes.xlsx", html)
        self.assertIn('id="question-import-file"', html)
        self.assertIn("/api/company/question-imports", script)
        self.assertIn("await saveExam('draft', true)", script)


if __name__ == "__main__":
    unittest.main()
