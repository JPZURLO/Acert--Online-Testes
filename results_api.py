import json
from io import BytesIO
from collections import defaultdict
from datetime import datetime, timedelta

from flask import Blueprint, jsonify, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from grading import grade_for_score


def clean_text(value, maximum):
    return str(value or "").strip()[:maximum]


def parse_json(value, default):
    try:
        parsed = json.loads(value or "")
        return parsed if isinstance(parsed, type(default)) else default
    except (TypeError, json.JSONDecodeError):
        return default


def result_label(score, passing_score, stored_status=None):
    if stored_status in {"approved", "review", "failed", "invalidated"}:
        return stored_status
    score = float(score or 0)
    passing_score = float(60 if passing_score is None else passing_score)
    if score >= passing_score:
        return "approved"
    if score >= max(0, passing_score - 10):
        return "review"
    return "failed"


def result_from_row(row, include_details=False):
    score = float(row.get("score") or 0)
    passing_score = float(60 if row.get("passing_score") is None else row["passing_score"])
    result = {
        "id": row["id"],
        "attemptId": row.get("attempt_id"),
        "identityStatus": row.get("identity_status") or "not_required",
        "participantId": row["participant_id"],
        "participantName": row.get("participant_name") or "Participante",
        "participantEmail": row.get("participant_email") or "",
        "examId": row["exam_id"],
        "examTitle": row.get("exam_title") or "Teste",
        "score": round(score, 2),
        "maxScore": float(row.get("max_score") or 100),
        "grade": grade_for_score(score, row.get("grading_scale_json")),
        "passingScore": round(passing_score, 2),
        "result": result_label(score, passing_score, row.get("result_status")),
        "durationSeconds": int(row.get("duration_seconds") or 0),
        "correctAnswers": int(row.get("correct_answers") or 0),
        "totalQuestions": int(row.get("total_questions") or 0),
        "completedAt": row.get("completed_at").isoformat() if row.get("completed_at") else None,
        "releaseStatus": row.get("release_status") or "released",
        "incidentCount": int(row.get("incident_count") or 0),
        "recordingStatus": row.get("recording_status") or "not_required",
    }
    if include_details:
        result["answers"] = parse_json(row.get("answers_json"), [])
        result["competencies"] = parse_json(row.get("competency_scores_json"), {})
        result["reviewerNotes"] = row.get("reviewer_notes") or ""
    return result


def format_excel_number(value):
    try:
        number = float(value)
    except (TypeError, ValueError):
        return 0
    return int(number) if number.is_integer() else round(number, 2)


def split_name(full_name):
    parts = str(full_name or "").strip().split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return "", parts[0]
    return " ".join(parts[1:]), parts[0]


def format_excel_datetime(value):
    if not value:
        return "-"
    if hasattr(value, "strftime"):
        return value.strftime("%d/%m/%Y %H:%M")
    return str(value)


def export_score_data(row):
    answers = parse_json(row.get("answers_json"), [])
    if not answers:
        return [], 0, 0
    earned = []
    total = 0.0
    score = 0.0
    for answer in answers:
        if not isinstance(answer, dict):
            continue
        points = max(0.0, float(answer.get("points") or 0))
        value = answer.get("earnedPoints")
        if value in (None, ""):
            earned_value = "-"
        else:
            earned_value = format_excel_number(max(0.0, min(points, float(value or 0))))
            score += float(earned_value)
        total += points
        earned.append({"points": format_excel_number(points), "earned": earned_value})
    return earned, format_excel_number(score), format_excel_number(total)


def compute_dashboard(rows):
    completed = len(rows)
    scores = [float(row.get("score") or 0) for row in rows]
    durations = [int(row.get("duration_seconds") or 0) for row in rows]
    labels = [result_label(row.get("score"), row.get("passing_score"), row.get("result_status")) for row in rows]
    approved = labels.count("approved")

    distribution = {
        "approved": approved,
        "review": labels.count("review"),
        "failed": labels.count("failed"),
    }

    competency_values = defaultdict(list)
    for row in rows:
        for name, value in parse_json(row.get("competency_scores_json"), {}).items():
            try:
                competency_values[clean_text(name, 80)].append(max(0, min(100, float(value))))
            except (TypeError, ValueError):
                continue
    competencies = [
        {"name": name, "score": round(sum(values) / len(values))}
        for name, values in competency_values.items()
        if name and values
    ]
    competencies.sort(key=lambda item: item["score"], reverse=True)

    today = datetime.now().date()
    weeks = []
    for offset in range(5, -1, -1):
        end = today - timedelta(days=offset * 7)
        start = end - timedelta(days=6)
        week_scores = []
        for row in rows:
            completed_at = row.get("completed_at")
            if completed_at and start <= completed_at.date() <= end:
                week_scores.append(float(row.get("score") or 0))
        weeks.append(
            {
                "label": f"{start.strftime('%d/%m')}–{end.strftime('%d/%m')}",
                "score": round(sum(week_scores) / len(week_scores), 1) if week_scores else 0,
            }
        )

    return {
        "stats": {
            "completed": completed,
            "averageScore": round(sum(scores) / completed, 1) if completed else 0,
            "approvalRate": round(approved / completed * 100) if completed else 0,
            "averageMinutes": round(sum(durations) / completed / 60) if completed else 0,
        },
        "distribution": distribution,
        "competencies": competencies[:6],
        "trend": weeks,
    }


def create_results_blueprint(open_database, token_payload):
    blueprint = Blueprint("company_results", __name__)

    def company_id_or_error():
        payload, error = token_payload("company")
        if error:
            return None, error
        try:
            return int(payload["sub"]), None
        except (KeyError, TypeError, ValueError):
            return None, (jsonify({"success": False, "message": "Sessão inválida."}), 401)

    @blueprint.get("/api/company/results")
    def list_results():
        company_id, error = company_id_or_error()
        if error:
            return error
        search = clean_text(request.args.get("search"), 180)
        exam_id = request.args.get("examId")
        status = clean_text(request.args.get("status"), 16)
        try:
            days = max(1, min(3650, int(request.args.get("days", 30))))
        except (TypeError, ValueError):
            days = 30

        cutoff = datetime.now() - timedelta(days=days)
        where = ["r.company_id = %s", "r.completed_at >= %s"]
        params = [company_id, cutoff]
        if search:
            term = f"%{search}%"
            where.append("(p.full_name LIKE %s OR p.email LIKE %s)")
            params.extend([term, term])
        if exam_id and str(exam_id).isdigit():
            where.append("r.exam_id = %s")
            params.append(int(exam_id))

        connection = open_database()
        cursor = connection.cursor(dictionary=True)
        try:
            cursor.execute("SELECT RazaoSocial FROM empresas WHERE id = %s", (company_id,))
            company = cursor.fetchone()
            sql = (
                "SELECT r.*, p.full_name AS participant_name, p.email AS participant_email, "
                "e.title AS exam_title, e.passing_score, e.grading_scale_json, a.identity_status, "
                "(SELECT COUNT(*) FROM attempt_audit_events ae WHERE ae.attempt_id=a.id AND ae.severity IN ('warning','critical')) AS incident_count, "
                "(SELECT ar.status FROM attempt_recordings ar WHERE ar.attempt_id=a.id LIMIT 1) AS recording_status FROM company_results r "
                "LEFT JOIN exam_attempts a ON a.id = r.attempt_id JOIN company_participants p ON p.id = r.participant_id AND p.company_id = r.company_id "
                "JOIN company_exams e ON e.id = r.exam_id AND e.company_id = r.company_id "
                f"WHERE {' AND '.join(where)} ORDER BY r.completed_at DESC LIMIT 2000"
            )
            cursor.execute(sql, tuple(params))
            rows = cursor.fetchall()
            if status in {"approved", "review", "failed", "invalidated"}:
                rows = [row for row in rows if result_label(row.get("score"), row.get("passing_score"), row.get("result_status")) == status]
            dashboard = compute_dashboard(rows)
            cursor.execute("SELECT id, title FROM company_exams WHERE company_id = %s ORDER BY title", (company_id,))
            exams = cursor.fetchall()
            return jsonify(
                {
                    "company": {"id": company_id, "name": company["RazaoSocial"] if company else "Empresa"},
                    "results": [result_from_row(row) for row in rows],
                    "exams": exams,
                    **dashboard,
                }
            )
        finally:
            cursor.close()
            connection.close()

    @blueprint.get("/api/company/results/<int:result_id>")
    def get_result(result_id):
        company_id, error = company_id_or_error()
        if error:
            return error
        connection = open_database()
        cursor = connection.cursor(dictionary=True)
        try:
            cursor.execute(
                "SELECT r.*, p.full_name AS participant_name, p.email AS participant_email, "
                "e.title AS exam_title, e.passing_score, e.grading_scale_json, a.identity_status, "
                "(SELECT COUNT(*) FROM attempt_audit_events ae WHERE ae.attempt_id=a.id AND ae.severity IN ('warning','critical')) AS incident_count, "
                "(SELECT ar.status FROM attempt_recordings ar WHERE ar.attempt_id=a.id LIMIT 1) AS recording_status FROM company_results r "
                "LEFT JOIN exam_attempts a ON a.id = r.attempt_id JOIN company_participants p ON p.id = r.participant_id AND p.company_id = r.company_id "
                "JOIN company_exams e ON e.id = r.exam_id AND e.company_id = r.company_id "
                "WHERE r.id = %s AND r.company_id = %s",
                (result_id, company_id),
            )
            row = cursor.fetchone()
            if not row:
                return jsonify({"success": False, "message": "Resultado não encontrado."}), 404
            result = result_from_row(row, include_details=True)
            attempt_id = row.get("attempt_id")
            result["auditEvents"] = []
            result["recording"] = None
            if attempt_id:
                cursor.execute(
                    "SELECT event_type,severity,details_json,occurred_at FROM attempt_audit_events "
                    "WHERE attempt_id=%s ORDER BY occurred_at,id",
                    (attempt_id,),
                )
                result["auditEvents"] = [
                    {
                        "type": event["event_type"],
                        "severity": event["severity"],
                        "details": parse_json(event.get("details_json"), {}),
                        "occurredAt": event["occurred_at"].isoformat() if event.get("occurred_at") else None,
                    }
                    for event in cursor.fetchall()
                ]
                cursor.execute(
                    "SELECT status,content_type,size_bytes,chunk_count,sha256,started_at,completed_at,available_until,delete_after,downloaded_at,deleted_at,deletion_reason "
                    "FROM attempt_recordings WHERE attempt_id=%s",
                    (attempt_id,),
                )
                recording = cursor.fetchone()
                if recording:
                    result["recording"] = {
                        "status": recording["status"],
                        "contentType": recording["content_type"],
                        "sizeBytes": int(recording.get("size_bytes") or 0),
                        "chunkCount": int(recording.get("chunk_count") or 0),
                        "sha256": recording.get("sha256"),
                        "startedAt": recording["started_at"].isoformat() if recording.get("started_at") else None,
                        "completedAt": recording["completed_at"].isoformat() if recording.get("completed_at") else None,
                        "availableUntil": recording["available_until"].isoformat() if recording.get("available_until") else None,
                        "deleteAfter": recording["delete_after"].isoformat() if recording.get("delete_after") else None,
                        "downloadedAt": recording["downloaded_at"].isoformat() if recording.get("downloaded_at") else None,
                        "deletedAt": recording["deleted_at"].isoformat() if recording.get("deleted_at") else None,
                        "deletionReason": recording.get("deletion_reason") or "",
                        "url": f"/api/company/attempts/{attempt_id}/recording" if recording.get("status") == "completed" else None,
                    }
            return jsonify({"result": result})
        finally:
            cursor.close()
            connection.close()

    @blueprint.get("/api/company/results/export.xlsx")
    def export_results_xlsx():
        company_id, error = company_id_or_error()
        if error:
            return error
        search = clean_text(request.args.get("search"), 180)
        exam_id = request.args.get("examId")
        status = clean_text(request.args.get("status"), 16)
        result_id = request.args.get("resultId")
        try:
            days = max(1, min(3650, int(request.args.get("days", 30))))
        except (TypeError, ValueError):
            days = 30

        cutoff = datetime.now() - timedelta(days=days)
        where = ["r.company_id = %s", "r.completed_at >= %s"]
        params = [company_id, cutoff]
        if result_id and str(result_id).isdigit():
            where.append("r.id = %s")
            params.append(int(result_id))
        if search:
            term = f"%{search}%"
            where.append("(p.full_name LIKE %s OR p.email LIKE %s)")
            params.extend([term, term])
        if exam_id and str(exam_id).isdigit():
            where.append("r.exam_id = %s")
            params.append(int(exam_id))

        connection = open_database()
        cursor = connection.cursor(dictionary=True)
        try:
            cursor.execute("SELECT RazaoSocial FROM empresas WHERE id = %s", (company_id,))
            company = cursor.fetchone()
            sql = (
                "SELECT r.*, p.full_name AS participant_name, p.email AS participant_email, p.phone AS participant_phone, p.city AS participant_city, "
                "e.title AS exam_title, e.passing_score, e.grading_scale_json, e.questions_json, a.started_at AS attempt_started_at "
                "FROM company_results r "
                "LEFT JOIN exam_attempts a ON a.id = r.attempt_id "
                "JOIN company_participants p ON p.id = r.participant_id AND p.company_id = r.company_id "
                "JOIN company_exams e ON e.id = r.exam_id AND e.company_id = r.company_id "
                f"WHERE {' AND '.join(where)} ORDER BY r.completed_at DESC LIMIT 3000"
            )
            cursor.execute(sql, tuple(params))
            rows = cursor.fetchall()
            if status in {"approved", "review", "failed", "invalidated"}:
                rows = [row for row in rows if result_label(row.get("score"), row.get("passing_score"), row.get("result_status")) == status]

            score_rows = [export_score_data(row) for row in rows]
            max_questions = max((len(item[0]) for item in score_rows), default=0)
            total_points = next((item[2] for item in score_rows if item[2]), 100)
            headers = [
                "Sobrenome",
                "Nome",
                "Endereço de email",
                "Telefone celular",
                "Cidade/Município",
                "Estado",
                "Iniciado em",
                "Completo",
                "Tempo utilizado",
                f"Avaliar/{str(total_points).replace('.', ',')}",
            ]
            for index in range(max_questions):
                points = next((items[0][index]["points"] for items in score_rows if len(items[0]) > index), 0)
                headers.append(f"Q. {index + 1} /{str(points).replace('.', ',')}")

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Resultados"
            sheet.append([company["RazaoSocial"] if company else "Online Teste"])
            sheet.append([rows[0]["exam_title"] if rows else "Resultados"])
            sheet.append([])
            sheet.append(headers)
            header_fill = PatternFill("solid", fgColor="E8F4F3")
            for cell in sheet[4]:
                cell.font = Font(bold=True, color="172033")
                cell.fill = header_fill

            question_totals = [[] for _ in range(max_questions)]
            score_values = []
            for row, score_data in zip(rows, score_rows):
                question_scores, raw_score, _raw_total = score_data
                surname, first_name = split_name(row.get("participant_name"))
                score_values.append(float(raw_score or 0))
                line = [
                    surname,
                    first_name,
                    row.get("participant_email") or "",
                    row.get("participant_phone") or "",
                    row.get("participant_city") or "",
                    "Finalizada" if row.get("completed_at") else "Nunca enviada",
                    format_excel_datetime(row.get("started_at") or row.get("attempt_started_at")),
                    format_excel_datetime(row.get("completed_at")),
                    f"{round(int(row.get('duration_seconds') or 0) / 60)} min",
                    raw_score,
                ]
                for index in range(max_questions):
                    value = question_scores[index]["earned"] if len(question_scores) > index else "-"
                    if value != "-":
                        question_totals[index].append(float(value))
                    line.append(value)
                sheet.append(line)

            if rows:
                average_line = ["Média geral", "", "", "", "", "", "", "", "", format_excel_number(sum(score_values) / len(score_values))]
                for values in question_totals:
                    average_line.append(format_excel_number(sum(values) / len(values)) if values else "-")
                sheet.append(average_line)
                for cell in sheet[sheet.max_row]:
                    cell.font = Font(bold=True)

            for column_cells in sheet.columns:
                length = max(len(str(cell.value or "")) for cell in column_cells)
                sheet.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 34)

            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            filename = "resultado-individual.xlsx" if result_id else "resultados-online-teste.xlsx"
            return send_file(
                output,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=filename,
            )
        finally:
            cursor.close()
            connection.close()

    @blueprint.put("/api/company/results/<int:result_id>/review")
    def review_result(result_id):
        company_id, error = company_id_or_error()
        if error:
            return error
        data = request.get_json(silent=True) or {}
        manual_scores = data.get("manualScores") if isinstance(data.get("manualScores"), dict) else {}
        notes = clean_text(data.get("notes"), 5000)
        release = bool(data.get("release"))
        requested_status = clean_text(data.get("resultStatus"), 24)
        connection = open_database()
        cursor = connection.cursor(dictionary=True)
        try:
            cursor.execute("SELECT r.*, e.passing_score, e.grading_scale_json FROM company_results r JOIN company_exams e ON e.id=r.exam_id AND e.company_id=r.company_id WHERE r.id=%s AND r.company_id=%s", (result_id, company_id))
            row = cursor.fetchone()
            if not row:
                return jsonify({"success": False, "message": "Resultado não encontrado."}), 404
            answers = parse_json(row.get("answers_json"), [])
            total_points = 0.0
            earned_points = 0.0
            manual_total = 0.0
            for answer in answers:
                if not isinstance(answer, dict):
                    continue
                points = max(0.0, float(answer.get("points") or 0))
                earned = max(0.0, min(points, float(answer.get("earnedPoints") or 0)))
                if answer.get("type") in {"long_answer", "essay"}:
                    value = manual_scores.get(str(answer.get("questionId") or ""), earned)
                    try:
                        earned = max(0.0, min(points, float(value)))
                    except (TypeError, ValueError):
                        earned = 0.0
                    answer["earnedPoints"] = round(earned, 2)
                    answer["isCorrect"] = earned >= points if points else None
                    answer["correctionStatus"] = "corrigido" if release else "em_correcao"
                    feedback_val = data.get("feedback", {}).get(str(answer.get("questionId") or "")) if isinstance(data.get("feedback"), dict) else None
                    if feedback_val:
                        answer["feedback"] = clean_text(feedback_val, 2000)
                    manual_total += earned
                total_points += points
                earned_points += earned
            score = round((earned_points / total_points * 100) if total_points else 0, 2)
            result_status = "review"
            if release:
                result_status = requested_status if requested_status in {"approved", "failed", "invalidated"} else ("approved" if score >= float(row.get("passing_score") or 60) else "failed")
            release_status = "released" if release else "pending"
            cursor.execute("UPDATE company_results SET score=%s,answers_json=%s,result_status=%s,release_status=%s,reviewer_notes=%s WHERE id=%s AND company_id=%s", (score, json.dumps(answers, ensure_ascii=False), result_status, release_status, notes, result_id, company_id))
            if row.get("attempt_id"):
                cursor.execute("UPDATE exam_attempts SET manual_score=%s,final_score=%s,review_status=%s,reviewer_notes=%s,reviewed_at=NOW() WHERE id=%s AND company_id=%s", (manual_total, score, "completed" if release else "pending", notes, row["attempt_id"], company_id))
            connection.commit()
            cursor.execute("SELECT r.*, p.full_name AS participant_name, p.email AS participant_email, e.title AS exam_title, e.passing_score, e.grading_scale_json FROM company_results r JOIN company_participants p ON p.id=r.participant_id AND p.company_id=r.company_id JOIN company_exams e ON e.id=r.exam_id AND e.company_id=r.company_id WHERE r.id=%s AND r.company_id=%s", (result_id, company_id))
            return jsonify({"success": True, "result": result_from_row(cursor.fetchone(), include_details=True)})
        except (TypeError, ValueError):
            connection.rollback()
            return jsonify({"success": False, "message": "As pontuações informadas são inválidas."}), 400
        finally:
            cursor.close()
            connection.close()
    return blueprint
